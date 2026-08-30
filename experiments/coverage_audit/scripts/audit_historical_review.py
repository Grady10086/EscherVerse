#!/usr/bin/env python3
"""Audit recoverable historical review evidence without inferring missing stages."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


def normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def item_key(row: dict[str, object]) -> tuple[str, str]:
    return normalize(row.get("P")), normalize(row.get("Q"))


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--benchmark", type=Path, required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--archive", type=Path, required=True)
    parser.add_argument("--output-json", type=Path, required=True)
    parser.add_argument("--output-md", type=Path, required=True)
    args = parser.parse_args()

    benchmark = json.loads(args.benchmark.read_text(encoding="utf-8"))
    final_by_key: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    final_by_video: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in benchmark:
        final_by_key[item_key(row)].append(row)
        final_by_video[normalize(row["P"])].append(row)

    workbook = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = workbook.active
    status_counts: Counter[str] = Counter()
    parse_counts: Counter[str] = Counter()
    effective: list[dict[str, object]] = []
    all_original_keys: list[tuple[str, str]] = []
    for excel_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        status = normalize(values[3])
        status_counts[status or "blank"] += 1
        try:
            original = json.loads(values[1]) if values[1] else None
        except (json.JSONDecodeError, TypeError):
            original = None
            parse_counts["original_json_error"] += 1
        try:
            revised = json.loads(values[4]) if values[4] else None
        except (json.JSONDecodeError, TypeError):
            revised = None
            parse_counts["revised_json_error"] += 1
        if original:
            all_original_keys.append(item_key(original))

        chosen = None
        source = None
        if status == "正确，不需要修改" and original:
            chosen, source = original, "original"
        elif status == "错误，需要修改" and revised and normalize(revised.get("A")):
            chosen, source = revised, "revised"
        elif status == "错误，需要修改":
            parse_counts["modification_required_but_unusable"] += 1
        else:
            parse_counts["untrusted_status"] += 1
        if chosen:
            effective.append({"excel_row": excel_row, "record": chosen, "source": source})

    effective_key_counts = Counter(item_key(entry["record"]) for entry in effective)
    effective_answers: dict[tuple[str, str], set[str]] = defaultdict(set)
    for entry in effective:
        record = entry["record"]
        effective_answers[item_key(record)].add(normalize(record.get("A")).lower())
    strict_effective_keys = {key for key in effective_key_counts if key in final_by_key}
    stable_final_answer_keys = {
        key
        for key in strict_effective_keys
        if len(effective_answers[key]) == 1
        and next(iter(effective_answers[key]))
        == normalize(final_by_key[key][0].get("A")).lower()
    }

    with zipfile.ZipFile(args.archive) as archive:
        archive_rows = [
            json.loads(archive.read(name))
            for name in archive.namelist()
            if name.endswith("/data.json")
        ]
    archive_strict = [row for row in archive_rows if item_key(row) in final_by_key]
    archive_changed_answer = [
        row
        for row in archive_strict
        if normalize(row.get("A")).lower()
        != normalize(final_by_key[item_key(row)][0].get("A")).lower()
    ]
    archive_same_category = sum(
        normalize(row.get("C")) == normalize(final_by_key[item_key(row)][0].get("C"))
        for row in archive_strict
    )
    archive_same_scene = sum(
        normalize(row.get("scene_type"))
        == normalize(final_by_key[item_key(row)][0].get("scene_type"))
        for row in archive_strict
    )

    report = {
        "schema": "escher-audit-historical-review-audit-v1",
        "sources": {
            "benchmark": {"path": str(args.benchmark), "sha256": sha256(args.benchmark)},
            "workbook": {"path": str(args.workbook), "sha256": sha256(args.workbook)},
            "archive": {"path": str(args.archive), "sha256": sha256(args.archive)},
        },
        "final_benchmark": {
            "rows": len(benchmark),
            "unique_video_question_keys": len(final_by_key),
            "unique_videos": len(final_by_video),
        },
        "manuscript_reported_video_attrition_not_log_verified": {
            "initial_online_videos": 3766053,
            "after_rule_filter": ">105000",
            "final_selected_videos": 11328,
            "stage_1_retention_lower_bound": 105000 / 3766053,
            "stage_1_removal_upper_bound": 1 - 105000 / 3766053,
            "stage_2_retention_upper_bound": 11328 / 105000,
            "stage_2_removal_lower_bound": 1 - 11328 / 105000,
            "reported_reason_taxonomy": {
                "rule_filter": ["static or monotonous video captions", "keyword/content-score mismatch"],
                "llm_filter": ["static or ambiguous content", "low object/agent interaction", "low action or movement value"],
            },
            "source": "sn-article.tex lines 341-347",
            "verification_status": "manuscript claims found; underlying stage logs not recovered",
        },
        "human_delivery_workbook": {
            "data_rows": sheet.max_row - 1,
            "status_counts": dict(sorted(status_counts.items())),
            "parse_and_usability_counts": dict(sorted(parse_counts.items())),
            "original_json_unique_keys": len(set(all_original_keys)),
            "effective_decision_rows": len(effective),
            "effective_original_rows": sum(entry["source"] == "original" for entry in effective),
            "effective_revised_rows": sum(entry["source"] == "revised" for entry in effective),
            "effective_unique_keys": len(effective_key_counts),
            "effective_unique_videos": len(
                {item_key(entry["record"])[0] for entry in effective}
            ),
            "repeated_effective_keys": sum(value > 1 for value in effective_key_counts.values()),
            "rows_in_repeated_effective_keys": sum(
                value for value in effective_key_counts.values() if value > 1
            ),
            "conflicting_effective_answer_keys": sum(
                len(answers) > 1 for answers in effective_answers.values()
            ),
            "unique_keys_strictly_mapped_to_final": len(strict_effective_keys),
            "unique_keys_with_one_human_answer_equal_to_final": len(stable_final_answer_keys),
        },
        "manual_review_archive_771": {
            "rows": len(archive_rows),
            "unique_keys": len({item_key(row) for row in archive_rows}),
            "unique_videos": len({normalize(row.get("P")) for row in archive_rows}),
            "strictly_mapped_to_final": len(archive_strict),
            "same_answer_as_final": len(archive_strict) - len(archive_changed_answer),
            "different_answer_from_final": len(archive_changed_answer),
            "same_category_as_final": archive_same_category,
            "same_scene_type_as_final": archive_same_scene,
            "interpretation": "A human-review handoff subset, not a candidate-generation or discard log.",
        },
        "recoverability": {
            "directly_observed": [
                "final benchmark size",
                "workbook review-row statuses and usable effective decisions",
                "archive-to-final answer/category/scene changes",
            ],
            "not_recoverable_from_these_artifacts": [
                "number of QA candidates initially generated",
                "number rejected before human handoff",
                "stage-specific discard reasons for the full pipeline",
                "generator-omitted question types",
            ],
            "prohibited_inference": "Do not subtract these non-nested artifacts to estimate attrition.",
        },
    }
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    w = report["human_delivery_workbook"]
    z = report["manual_review_archive_771"]
    video = report["manuscript_reported_video_attrition_not_log_verified"]
    markdown = f"""# E5a 历史审核与流失证据审计

## 稿件声明的视频筛选规模（未由原始日志独立复算）

| 阶段 | 输入 | 保留 | 可报告边界 | 稿件所述主要排除依据 |
|---|---:|---:|---:|---|
| Rule-based caption filter | 3,766,053 | 105,000+ | 保留率 >={video['stage_1_retention_lower_bound']:.2%}；移除率 <={video['stage_1_removal_upper_bound']:.2%} | static/monotonous；关键词与内容分值不足 |
| LLM-based video filter | 105,000+ | 11,328 | 保留率 <={video['stage_2_retention_upper_bound']:.2%}；移除率 >={video['stage_2_removal_lower_bound']:.2%} | static/ambiguous；对象/主体交互、动作或运动价值不足 |

以上数字来自稿件 `sn-article.tex` 的流程描述；尚未恢复原始逐视频阶段日志，因此不能给出精确的 105,000 后输入量、逐原因数量或独立复算结果。

## 可直接观测的规模

| 工件/口径 | 数量 |
|---|---:|
| 最终 benchmark 行数 | {len(benchmark):,} |
| 人工交付表记录行 | {sheet.max_row - 1:,} |
| 人工交付表：正确、不需修改 | {status_counts['正确，不需要修改']:,} |
| 人工交付表：错误、需要修改 | {status_counts['错误，需要修改']:,} |
| 可形成有效人工结论的记录行 | {w['effective_decision_rows']:,} |
| 有效人工结论的唯一视频+题干 | {w['effective_unique_keys']:,} |
| 严格映射至当前最终 benchmark 的唯一题干 | {w['unique_keys_strictly_mapped_to_final']:,} |
| 重复交付的有效题干 | {w['repeated_effective_keys']:,} |
| 出现相互冲突人工答案的题干 | {w['conflicting_effective_answer_keys']:,} |
| 771 题历史人工审核包：严格映射至最终题 | {z['strictly_mapped_to_final']:,} |
| 其中答案与最终版本相同 | {z['same_answer_as_final']:,} |
| 其中答案在后续发生变化 | {z['different_answer_from_final']:,} |

## 解释边界

这些工件证明了人工审核、修改与冲突的存在，但并非彼此严格嵌套的生成阶段日志。特别是 771 题压缩包是人工交付子集，3,569 行表格含重复交付、冲突和未完成修改，不能用二者相减恢复候选淘汰率。

现有工件无法直接恢复：初始 QA 候选总量、进入人工前的淘汰量、全流程逐阶段 discard reason，以及生成器从未提出的问题类型。E5b 的独立盲化参考集专门测量最后一项；其余缺失数字在回复中应明确标为历史日志不可恢复，而非倒推估计。
"""
    args.output_md.write_text(markdown, encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
